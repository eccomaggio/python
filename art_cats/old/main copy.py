from openpyxl import load_workbook
from dataclasses import dataclass
from collections import namedtuple
from typing import List
from enum import Enum, auto
from pprint import pprint
from datetime import datetime, timezone
import re
from pathlib import Path
import sys

@dataclass
class Title:
    original: str
    transliteration: str


@dataclass
class Record:
    sublib: str
    lang: List[str]
    isbn: str
    title: Title
    subtitle: Title
    parallel_title: Title
    parallel_subtitle: Title
    country: str
    place: str
    publisher: str
    pub_year: str
    copyright: str
    extent: int
    size: int
    series_title: str
    series_enum: str
    notes: str
    sales_code: str
    sale_dates: List[str]
    hol_note: str
    donation: str
    barcode: str
    pub_year_is_approx: bool
    extent_is_approx: bool
    timestamp: datetime
    sequence_number: int
    links: List


def norm_langs(raw):
    lang_list = {
        "english": "eng",
        "chinese": "chi",
        "german": "ger",
        "italian": "ita",
        "spanish": "spa",
        "french": "fre",
        "swedish": "swe",
        "danish": "dan",
        "norwegian": "nor",
    }
    result = []
    langs = raw.replace(" ", "").lower().split("/")
    for lang in langs:
        try:
            result.append(lang_list[lang])
        except KeyError as e:
            print(f"Warning: {e} is not a recognised language; it has been passed on unchanged.")
            result.append(lang)
    # result = [lang_list[lang.strip().lower()] for lang in raw.split("/")]
    return result

def norm_country(country):
    countries = {
        "china": "cc",
        "france": "fr",
        "germany": "gw",
        "italy": "it",
        "portugal": "po",
        "netherlands": "ne",
        "spain": "sp",
        "sweden": "sw",
        "denmark": "dk",
        "norway": "no",
        "canada": "xxc",
        "new york": "nyu",
        "england": "enk",
        "uk": "xxk",
        "us": "xxu",
        "usa": "xxu",
        "austria": "au"
    }
    try:
        result = countries[country.strip().lower()]
    except KeyError as e:
        print(f"Warning: {e} is not a recognised country name; it has been passed on unchanged.")
        result = country
    return result

def norm_dates(raw):
    result = [date.strip() for date in raw.split(",")]
    return result


def normalize(entry):
    return str(entry).strip() if entry else None


def check_for_approx(raw):
    raw = trim_mistaken_decimals(str(raw).strip())
    # if raw.endswith(".0"):
    #     raw = raw[:-2]
    is_approx = False
    if raw[-1] == "?":
        is_approx = True
        raw = raw[:-1].rstrip()
    return (raw, is_approx)

def trim_mistaken_decimals(string):
    if not string:
        return
    if string.endswith(".0"):
        string = string[:-2]
    return string


def get_records(excel_file):
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    print(f"\n{sheet.title} in {excel_file}")
    current_time = datetime.now()
    records = []
    for value in sheet.iter_rows(min_row=2, values_only=True):
        if not value[0]:
            break
        sublibrary = normalize(value[0])
        langs = norm_langs(value[1])
        isbn = trim_mistaken_decimals(normalize(value[2]))
        title = Title(normalize(value[3]), normalize(value[4]))
        subtitle = Title(normalize(value[5]), normalize(value[6]))
        parallel_title = Title(normalize(value[7]), normalize(value[8]))
        parallel_subtitle = Title(normalize(value[9]), normalize(value[10]))
        country = norm_country(value[11])
        place = normalize(value[12])
        publisher = normalize(value[13])
        pub_date, pub_date_is_approx = check_for_approx(str(value[14]))
        copyright_ = trim_mistaken_decimals(normalize(value[15]))
        extent, extent_is_approx = check_for_approx(value[16])
        size = int(value[17])
        series_title = normalize(value[18])
        series_enum = trim_mistaken_decimals(normalize(value[19]))
        note = normalize(value[20])
        sale_code = trim_mistaken_decimals(normalize(value[21]))
        date_of_sale = normalize(value[22]).replace(r"\s","").replace(".0", "").split(",")
        hol_notes = normalize(value[23])
        donation = normalize(value[24])
        barcode = trim_mistaken_decimals(normalize(value[25]))
        sequence_number = 1
        links = [880, []]

        record = Record(
            sublibrary,
            langs,
            isbn,
            title,
            subtitle,
            parallel_title,
            parallel_subtitle,
            country,
            place,
            publisher,
            pub_date,
            copyright_,
            int(extent),
            size,
            series_title,
            series_enum,
            note,
            sale_code,
            date_of_sale,
            hol_notes,
            donation,
            barcode,

            pub_date_is_approx,
            extent_is_approx,
            current_time,
            sequence_number,
            links
        )
        records.append(record)
        pprint(record)
    return records


def build_000():
    """leader (0 is only for sorting purposes; should read 'LDR')"""
    return build_field(0, [[-2,-2,"00000nam a22000003i 4500"]])


def build_005(record):
    """date & time of transaction
    "The date requires 8 numeric characters in the pattern yyyymmdd. The time requires 8 numeric characters in the pattern hhmmss.f, expressed in terms of the 24-hour (00-23) clock."
    """
    i1, i2 = -2, -2  ## "This field has no indicators or subfield codes."
    standard_time = record.timestamp.now(timezone.utc)
    ## NB: python produces this format: YYYY-MM-DD HH:MM:SS.ffffff, e.g. 2020-09-30 12:37:55.713351
    timestamp = str(standard_time).translate(str.maketrans("", "", " -:"))[:16]
    return build_field(5, [[i1, i2, f"{timestamp}"]])


def build_008(record):
    """pub year & main language?"""
    i1, i2 = -2, -2  ## "Field has no indicators or subfield codes"
    t = record.timestamp
    date_entered_on_file = str(t.year)[2:] + str(t.month).zfill(2) + str(t.day).zfill(2)
    pub_status = "s"
    date_1 = record.pub_year
    date_2 = 4 * "|"
    place_of_pub = record.country.ljust(3, "\\")
    books_configuration = (14*"|") + "\\" + (2*"|")
    lang = record.lang[0].ljust(3, "\\")
    modified_and_cataloging = 2*"|"
    # return build_field( 8, [[-2,-2, f"{str(t.year)[2:]}{str(t.month).zfill(2)}{str(t.day).zfill(2)}s{record.pub_year}{4*"|"}{record.country}{14*"|"}\\||eng||" ]])
    return build_field( 8, [[i1, i2, f"{date_entered_on_file}{pub_status}{date_1}{date_2}{place_of_pub}{books_configuration}{lang}{modified_and_cataloging}" ]])


def build_033(record):
    """sales dates"""
    i1 = 0 if len(record.sale_dates) == 1 else 1
    i2 = -1
    date_string = f"$a{"$a".join(record.sale_dates)}"
    return build_field(33, [[i1, i2, date_string]])


def build_040():
    """catalogued in Oxford"""
    return build_field(40, [[-1,-1,"$aUkOxU$beng$erda$cUkOxU"]])


def build_245(record):
    """ Title """
    has_chinese_title = bool(record.title.transliteration)
    if has_chinese_title:
        title, subtitle = record.title.transliteration, record.subtitle.transliteration
        chinese_title = combine(record.title.original, record.subtitle.original)
        nonfiling = 0
        sequence_number = seq_num(record.sequence_number)
        linkage = f"$6880-{sequence_number}"
    else:
        title, subtitle = record.title.original, record.subtitle.original
        nonfiling, title = mark_nonfiling_words(title)
        linkage = ""
    title = combine(title, subtitle)
    title = end_title_with_period(title)
    # if title[-1] not in "?!.":
    #     title = title + "."
    i1, i2 = 0, nonfiling
    if has_chinese_title:
        build_880(record, chinese_title, i1, i2, "245", sequence_number)
    return build_field(245, [[i1, i2, f"{linkage}$a{title}"]])


def build_264(record):
    """publisher & copyright"""
    i1 = -1
    i2 = 1  ## "Publication: Field contains a statement relating to the publication, release, or issuing of a resource."
    result = []
    place = f"$a{record.place} "
    publisher = f":$b{record.publisher}"
    pub_year = record.pub_year
    if record.pub_year_is_approx:
        pub_year = f"[{pub_year}?]"
    pub_year = f",$c{pub_year}"
    result.append([i1, i2, f"{place}{publisher}{pub_year}"])

    if record.copyright:
        result.append([i1, i2, f"$a\u00a9 {record.copyright}"])
    return build_field(264, result)


def build_300(record):
    """physical description"""
    i1, i2 = -1, -1 ## "undefined"
    pages = f"{record.extent} pages"
    if record.extent_is_approx:
        pages = f"approximately {pages}"
    size = f"{record.size} cm"
    return build_field(300, [[i1, i2, f"$a{pages} ;$c{size}"]])


def build_336():
    """content type (boilerplate)"""
    return build_field(336, [[-1, -1, "$atext$2rdacontent"]])


def build_337():
    """media type (boilerplate)"""
    return build_field(337, [[-1, -1, "$aunmediated$2rdamedia"]])


def build_338():
    """carrier type (boilerplate)"""
    return build_field(338, [[-1, -1, "$avolume$2rdacarrier"]])


def build_876(record):
    """notes / donations / barcode"""
    notes = f"$z{record.notes}" if record.notes else ""
    donation = f"$z{record.donation}" if record.donation else ""
    barcode = f"$p{record.barcode}"
    return build_field(876, [[-1, -1, f"{barcode}{donation}{notes}"]])


def build_904():
    """authority (boilerplate)"""
    return build_field(904, [[-1, -1, "$aOxford Local Record"]])


def build_024(record):  ##optional
    """sales code (if exists)"""
    result = (
        build_field(24, [[8, -1, f"$a{record.sales_code.strip()}"]])
        if record.sales_code
        else ""
    )
    return result


def build_041(record):  ##optional
    """language codes if not monolingual"""
    i1 = (
        -1
    )  ## "No information is provided as to whether the item is or includes a translation."
    i2 = -1  ## "(followed by) MARC language code"
    is_multi_lingual = len(record.lang) > 1
    if is_multi_lingual:
        if record.title.transliteration:
            main_lang = record.lang[0]
            others = record.lang[1:]
            lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        else:
            lang_list = f"$a{"$a".join(record.lang)}"
        # result = (build_field(41, [[i1,i2, f"$a{"$a".join(others)}$h{main_lang}"]]))
        result = (build_field(41, [[i1,i2, lang_list]]))
    else:
        result = ""
    return result


def build_246(record):  ##optional
    """
    Varying Form of Title
    Holds parallel Western title AND/OR original Chinese character title
    NB: Initial articles (e.g., The, La) are generally not recorded in field 246 unless the intent is to file on the article. [https://www.loc.gov/marc/bibliographic/bd246.html]
    """
    i1 = 3  ## "No note, added entry"
    i2 = 1  ## parallel title
    result = ""
    has_parallel_title = bool(record.parallel_title.original)
    # has_chinese_main_title = bool(record.title.transliteration)
    has_chinese_parallel_title = bool(record.parallel_title.transliteration)
    linkage = "$6880-01" if has_chinese_parallel_title else ""
    if has_chinese_parallel_title:
        parallel_title = record.parallel_title.transliteration
        parallel_subtitle = record.parallel_subtitle.transliteration
        chinese_parallel_title = combine(record.parallel_title.original, record.parallel_subtitle.original)
        sequence_number = seq_num(record.sequence_number)
        linkage = f"$6880-{sequence_number}"
        build_880(record, chinese_parallel_title,i1, i2, "246", sequence_number)

    elif has_parallel_title:  ## (i.e. Western script)
        parallel_title, parallel_subtitle = record.parallel_title.original, record.parallel_subtitle.original
        # nonfiling, parallel_title = mark_nonfiling_words(parallel_title)

    else:
        parallel_title, parallel_subtitle = "", ""
    if parallel_title:
        parallel_title = combine(parallel_title, parallel_subtitle)

    # ## Combine Chinese character main title with Western parallel title (if either exists)
    # ## (not sure how to handle Chinese original title + parallel titles)
    # if has_chinese_main_title:
    #     result = combine(chinese_character_title, parallel_title)
    # else:
    #     result = parallel_title

    # result = build_field(246, [[0, nonfiling, f"$a{result}"]]) if result else ""
    result = build_field(246, [[i1, i2, f"{linkage}$a{parallel_title}"]]) if parallel_title else ""
    return result

def build_490(record):  ## optional
    """Series Statement"""
    i1 = 0  ## Series not traced: No series added entry is desired for the series.
    i2 = -1
    series_title = f"$a{record.series_title}" if record.series_title else ""
    series_enum = f"$v{record.series_enum}" if record.series_enum else ""
    sep = " ;" if series_title and series_enum else ""
    content = series_title + sep + series_enum
    result = build_field(490, [[i1, i2, content ]]) if content else ""
    return result


def build_500(record):  ##optional
    """general notes"""
    result = (
        build_field(500, [[-1, -1, f"$a{record.hol_note}"]]) if record.hol_note else ""
    )
    return result

def build_880(record, title, i1, i2, caller, sequence_number):  ##optional
    """Alternate Graphic Representation"""
    record.sequence_number += 1
    # record.links[1].append(build_field(880, [[i1, i2, f"$6{caller}-{sequence_number}$a{title}"]]))
    record.links[1].append(build_line(line_prefix("880"), i1, i2, f"$6{caller}-{sequence_number}$a{title}"))


def build_field(numeric_tag, data):
    if data:
        lines = []
        display_tag = "LDR" if numeric_tag == 0 else seq_num(numeric_tag)
        # line_start = f"{display_tag}=  "
        for line in data:
            i1 = expand_indicators(line[0])
            i2 = expand_indicators(line[1])
            content = line[2]
            # lines.append(f"{line_start}{i1}{i2}{content}")
            lines.append(build_line(line_prefix(numeric_tag), i1, i2, content))
        result = (numeric_tag, lines)
    else:
        result = ""
    return result

def build_line(line_start, i1, i2, content):
    return (f"{line_start}{i1}{i2}{content}")

def line_prefix(numeric_tag):
    display_tag = "LDR" if numeric_tag == 0 else seq_num(numeric_tag)
    return f"={field_prefix(display_tag)}  "

def field_prefix(field_number):
    return str(field_number).zfill(3)

def seq_num(sequence_number):
    return str(sequence_number).zfill(2)


def expand_indicators(indicator):
    expansions = {-2: "", -1: "\\"}
    return expansions.get(indicator, indicator)


def combine(title, subtitle, sep=" :$b"):
    if subtitle:
        title += sep + subtitle
    return title

def end_title_with_period(title):
    if title[-1] not in "?!.":
        title = title + "."
    return title


def mark_nonfiling_words(title):
    nonfiling = 0
    if title:
        test = title.split("@@")
        if len(test) > 1:
            nonfiling = len(test[0])
            title = test[0] + test[1]
    return (nonfiling, title)


def check_for_nonfiling(title, lang="eng"):
    nonfiling_chars = {
        "eng": ("the", "a", "an"),
        "fr": ("le", "la", "les", "l'", "un", "une"),
        "it": ("lo", "il", "i", "l'", "gli", "le", "un", "una", "un'"),
        "sp": ("el", "la", "las", "los", "un", "una", "unos", "unas"),
        "gw": ("der", "die", "das", "ein", "eine"),
        "ne": ("de", "het"),
        "sw": ("en", "ett", "den", "det", "de"),
        "dk": ("en", "et"),
        "no": ("en", "ei", "et"),
        "po": ("o", "a", "os", "as", "um", "uma", "uns", "umas"),
        "chi": (),
    }
    # lang = "eng"
    break_char = "@@"
    nonfiling = title.find(break_char)
    result = (0, title)
    if nonfiling > 0:
        result = (nonfiling, title.replace(break_char, "", 1))
    else:
        for article in nonfiling_chars[lang]:
            test = re.match(f"({article}\\s?[^\\w\\s]?)\\w", title, re.I)
            if test:
                result = (test.span()[1] - 1, title)
                break
    return result


def build_mark_records(records):
    mark_records = []
    for record in records:
        mark_record = []
        for boilerplate in (
            build_000,
            build_040,
            build_336,
            build_337,
            build_338,
            build_904,
        ):
            mark_record.append(boilerplate())

        for step in (
            build_005,
            build_008,
            build_033,
            build_245,
            build_264,
            build_300,
            build_490,
            build_876,
            build_024,
            build_041,
            build_246,
            build_500,
        ):
            field = step(record)
            if field:
                mark_record.append(field)
        if record.links:
            mark_record.append(record.links)
        mark_record.sort(key=lambda field: field[0])
        mark_records.append(mark_record)

    return mark_records

def write_mrk_file(data, file_name="output.mrk"):
    mrk_file_dir = Path("mrk_files")
    if not mrk_file_dir.is_dir():
        mrk_file_dir.mkdir()
        try:
            mrk_file_dir.mkdir()
            print(f"Directory '{mrk_file_dir}' created successfully.")
        except FileExistsError:
            print(f"Directory '{mrk_file_dir}' already exists.")
        except PermissionError:
            print(f"Permission denied: Unable to create '{mrk_file_dir}'.")
        except Exception as e:
            print(f"An error occurred: {e}")

    out_file = mrk_file_dir / file_name
    with open(out_file, "w") as f:
        for record in data:
            for field in record:
                for line in field[1]:
                    # print(line)
                    f.write(line)
                    f.write("\n")
            f.write("\n")


def main():
    # excel_file = "reviews-sample.xlsx"ChineseArtCatalogues.3examples.corrected
    # excel_file = "sample.xlsx"
    # excel_file = "excel_files/ChineseArtCatalogues.3examples.updated.xlsx"
    excel_file = "ChineseArtCatalogues.3examples.updated.xlsx"
    # excel_file = "artCats.ernstHauswedell.xlsx"
    # excel_file = "auction_catalogues.PTW17_01_2025.xlsx"
    excel_file_address = Path("excel_files") / excel_file
    records = get_records(str(excel_file_address.resolve()))
    # pprint(records)
    mark_records = build_mark_records(records)
    write_mrk_file(mark_records)

    pprint(records)

if __name__ == "__main__":
    main()
