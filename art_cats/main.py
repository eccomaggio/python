from openpyxl import load_workbook  # type: ignore
from dataclasses import dataclass, fields
from typing import List, Any
# from pprint import pprint
from datetime import datetime, timezone
import re
from pathlib import Path
import logging


logger = logging.getLogger(__name__)
logging.basicConfig(
    filename="output.log",
    filemode="w",
    encoding="utf-8",
    format="%(levelname)s:%(message)s",
    level=logging.DEBUG
    )


@dataclass
class Title:
    original: str
    transliteration: str


@dataclass
class Record:
    sublib: str
    langs: List[str]
    isbn: str
    title: Title
    subtitle: Title
    parallel_title: Title
    parallel_subtitle: Title
    country: str
    place: str
    publisher: str
    pub_year: str
    copyright: str
    extent: str
    size: int
    series_title: str
    series_enum: str
    notes: str
    sales_code: str
    sale_dates: List[str]
    hol_notes: str
    donation: str
    barcode: str
    pub_year_is_approx: bool
    extent_is_approx: bool
    timestamp: datetime
    sequence_number: int
    links: List


@dataclass
class Result:
    is_ok:tuple[int, Any] | None    ## [field number, returned data]
    is_err:tuple[int, str] | None   ## [field number, error message]


class MissingFieldError(Exception):
    pass


def norm_langs(raw: str) -> list[str]:
    language_codes = {
        "english": "eng",
        "chinese": "chi",
        "german": "ger",
        "italian": "ita",
        "spanish": "spa",
        "french": "fre",
        "swedish": "swe",
        "danish": "dan",
        "norwegian": "nor",
        "dutch": "dut",
    }
    list_of_languages = []
    languages: List[str] = raw.replace(" ", "").lower().split("/")
    for language in languages:
        try:
            list_of_languages.append(language_codes[language])
        except KeyError as e:
            logger.warning(f"Warning: {e} is not a recognised language; it has been passed on unchanged.")
            list_of_languages.append(language)
    # result = [lang_list[lang.strip().lower()] for lang in raw.split("/")]
    return list_of_languages


def norm_country(country_raw: str) -> str:
    country_codes = {
        "algeria": "ae",
        "angola": "ao",
        "benin": "dm",
        "botswana": "bs",
        "burkinafaso": "uv",
        "burundi": "bd",
        "cameroon": "cm",
        "centralafricanrepublic": "cx",
        "chad": "cd",
        "congo": "cf",
        "democraticrepublicofcongo": "cg",
        "côtedivoire": "iv",
        "cotedivoire": "iv",
        "djibouti": "ft",
        "egypt": "ua",
        "equatorialguinea": "eg",
        "eritrea": "ea",
        "ethiopia": "et",
        "gabon": "go",
        "gambia": "gm",
        "ghana": "gh",
        "guinea": "gv",
        "guineabissau": "pg",
        "kenya": "ke",
        "lesotho": "lo",
        "liberia": "lb",
        "libya": "ly",
        "madagascar": "mg",
        "malawi": "mw",
        "mali": "ml",
        "mauritania": "mu",
        "morocco": "mr",
        "mozambique": "mz",
        "namibia": "sx",
        "niger": "ng",
        "nigeria": "nr",
        "rwanda": "rw",
        "saotomeandprincipe": "sf",
        "senegal": "sg",
        "sierraleone": "sl",
        "somalia": "so",
        "southafrica": "sa",
        "southsudan": "sd",
        "spanishnorthafrica": "sh",
        "sudan": "sj",
        "swaziland": "sq",
        "tanzania": "tz",
        "togo": "tg",
        "tunisia": "ti",
        "uganda": "ug",
        "westernsahara": "ss",
        "zambia": "za",
        "zimbabwe": "rh",
        "afghanistan": "af",
        "armenia": "ai",
        "republicofarmenia": "ar",
        "azerbaijan": "aj",
        "bahrain": "ba",
        "bangladesh": "bg",
        "bhutan": "bt",
        "brunei": "bx",
        "burma": "br",
        "cambodia": "cb",
        "china": "cc",
        "cyprus": "cy",
        "easttimor": "em",
        "gazastrip": "gz",
        "georgia": "gs",
        "georgianrepublic": "gs",
        "republicofgeorgia": "gs",
        "india": "ii",
        "indonesia": "io",
        "iran": "ir",
        "iraq": "iq",
        "israel": "is",
        "japan": "ja",
        "jordan": "jo",
        "kazakhstan": "kz",
        "northkorea": "kn",
        "korea": "ko",
        "southkorea": "ko",
        "kuwait": "ku",
        "kyrgyzstan": "kg",
        "laos": "ls",
        "lebanon": "le",
        "malaysia": "my",
        "mongolia": "mp",
        "nepal": "np",
        "oman": "mk",
        "pakistan": "pk",
        "papuanewguinea": "pp",
        "paracelislands": "pf",
        "philippines": "ph",
        "qatar": "qa",
        "saudiarabia": "su",
        "singapore": "si",
        "spratlyisland": "xp",
        "srilanka": "ce",
        "syria": "sy",
        "tajikistan": "ta",
        "thailand": "th",
        "turkey": "tu",
        "turkmenistan": "tk",
        "unitedarabemirates": "ts",
        "uae": "ts",
        "uzbekistan": "uz",
        "vietnam": "vm",
        "westbankofthejordanriver": "wj",
        "westbank": "wj",
        "yemen": "ye",
        "bermudaislands": "bm",
        "bermuda": "bm",
        "bouvetisland": "bv",
        "caboverde": "cv",
        "faroeislands": "fa",
        "faroes": "fa",
        "falklandislands": "fk",
        "falklands": "fk",
        "sainthelena": "xj",
        "southgeorgiaandthesouthsandwichislands": "xs",
        "southgeorgia": "xs",
        "southsandwichislands": "xs",
        "belize": "bh",
        "costarica": "cr",
        "elsalvador": "es",
        "guatemala": "gt",
        "honduras": "ho",
        "nicaragua": "nq",
        "panama": "pn",
        "albania": "aa",
        "andorra": "an",
        "austria": "au",
        "belarus": "bw",
        "belgium": "be",
        "bosniaandherzegovina": "bn",
        "bosnia": "bn",
        "bosniaherzegovina": "bn",
        "herzegovina": "bn",
        "bulgaria": "bu",
        "croatia": "ci",
        "czechrepublic": "xr",
        "czechia": "xr",
        "denmark": "dk",
        "estonia": "er",
        "finland": "fi",
        "france": "fr",
        "germany": "gw",
        "gibraltar": "gi",
        "greece": "gr",
        "guernsey": "gg",
        "hungary": "hu",
        "iceland": "ic",
        "ireland": "ie",
        "isleofman": "im",
        "italy": "it",
        "jersey": "je",
        "kosovo": "kv",
        "latvia": "lv",
        "liechtenstein": "lh",
        "lithuania": "li",
        "luxembourg": "lu",
        "macedonia": "xn",
        "malta": "mm",
        "montenegro": "mo",
        "moldova": "mv",
        "monaco": "mc",
        "netherlands": "ne",
        "norway": "no",
        "poland": "pl",
        "portugal": "po",
        "serbia": "rb",
        "romania": "rm",
        "russia": "ru",
        "russianfederation": "ru",
        "sanmarino": "sm",
        "slovakia": "xo",
        "slovenia": "xv",
        "spain": "sp",
        "sweden": "sw",
        "switzerland": "sz",
        "ukraine": "un",
        "vaticancity": "vc",
        "serbiaandmontenegro": "yu",
        "britishindianoceanterritory": "bi",
        "christmasisland": "xa",
        "cocosislands": "xb",
        "keelingislands": "xb",
        "comoros": "cq",
        "heardandmcdonaldislands": "hm",
        "maldives": "xc",
        "mauritius": "mf",
        "mayotte": "ot",
        "réunion": "re",
        "reunion": "re",
        "seychelles": "se",
        "americansamoa": "as",
        "cookislands": "cw",
        "fiji": "fj",
        "frenchpolynesia": "fp",
        "guam": "gu",
        "johnstonatoll": "ji",
        "kiribati": "gb",
        "marshallislands": "xe",
        "micronesia": "fm",
        "federatedstatesofmicronesia": "fm",
        "midwayislands": "xf",
        "nauru": "nu",
        "newcaledonia": "nl",
        "niue": "xh",
        "northernmarianaislands": "nw",
        "palau": "pw",
        "pitcairnisland": "pc",
        "samoa": "ws",
        "solomonislands": "bp",
        "tokelau": "tl",
        "tonga": "to",
        "tuvalu": "tv",
        "vanuatu": "nn",
        "wakeisland": "wk",
        "wallisandfutuna": "wf",
        "wallis": "wf",
        "futuna": "wf",
        "argentina": "ag",
        "bolivia": "bo",
        "brazil": "bl",
        "chile": "cl",
        "colombia": "ck",
        "ecuador": "ec",
        "frenchguiana": "fg",
        "guyana": "gy",
        "paraguay": "py",
        "peru": "pe",
        "surinam": "sr",
        "uruguay": "uy",
        "venezuela": "ve",
        "anguilla": "am",
        "antiguaandbarbuda": "aq",
        "antigua": "aq",
        "barbuda": "aq",
        "aruba": "aw",
        "bahamas": "bf",
        "barbados": "bb",
        "britishvirginislands": "vb",
        "caribbeannetherlands": "ca",
        "caymanislands": "cj",
        "cuba": "cu",
        "curaçao": "co",
        "curacao": "co",
        "dominica": "dq",
        "dominicanrepublic": "dr",
        "grenada": "gd",
        "guadeloupe": "gp",
        "haiti": "ht",
        "jamaica": "jm",
        "martinique": "mq",
        "montserrat": "mj",
        "puertorico": "pr",
        "saintbarthélemy": "sc",
        "saintbarthelemy": "sc",
        "saintkittsnevis": "xd",
        "saintkitts": "xd",
        "nevis": "xd",
        "saintlucia": "xk",
        "saintmartin": "st",
        "saintvincentandthegrenadines": "xm",
        "saintvincent": "xm",
        "thegrenadines": "xm",
        "grenadines": "xm",
        "sintmaarten": "sn",
        "trinidadandtobago": "tr",
        "trinidad": "tr",
        "tobago": "tr",
        "turksandcaicosislands": "tc",
        "virginislandsoftheunitedstates": "vi",
        "antarctica": "ay",
        "noplace": "xx",
        "unknown": "xx",
        "undetermined": "xx",
        "variousplaces": "vp",
        "various": "vp",
    }
    # normed_country = country.replace(" ", "").lower()
    country = re.sub(r"[\s\-']", "", country_raw).lower()
    try:
        result = country_codes[country]
    except KeyError as e:
        if len({e}) < 4:
            logger.info(f"Advisory: assuming country name ({e}) has already been processed.")
        else:
            logger.warning(f"Warning: {e} is not a recognised country name; it has been passed on unchanged.")
        result = country_raw
    return result


def norm_place(place_raw: str) -> str:
    long_country_codes = {
        "england": "enk",
        "northernireland": "nik",
        "scotland": "stk",
        "wales": "wlk",

        "alberta": "abc",
        "britishcolumbia": "bcc",
        "bc": "bcc",
        "manitoba": "mbc",
        "newbrunswick": "nkc",
        "newfoundland": "nfc",
        "labrador": "nfc",
        "newfoundlandandlabrador": "nfc",
        "northwestterritories": "ntc",
        "novascotia": "nsc",
        "nunavut": "nuc",
        "ontario": "onc",
        "princeedwardisland": "pic",
        "québecprovince": "quc",
        "quebéc": "quc",
        "quebecprovince": "quc",
        "quebec": "quc",
        "saskatchewan": "snc",
        "yukonterritory": "ykc",
        "yukon": "ykc",

        "alabama": "alu",
        "alaska": "aku",
        "arizona": "azu",
        "arkansas": "aru",
        "california": "cau",
        "colorado": "cou",
        "connecticut": "ctu",
        "delaware": "deu",
        "districtofcolumbia": "dcu",
        "columbia": "dcu",
        "florida": "flu",
        "georgia": "gau",
        "hawaii": "hiu",
        "idaho": "idu",
        "illinois": "ilu",
        "indiana": "inu",
        "iowa": "iau",
        "kansas": "ksu",
        "kentucky": "kyu",
        "louisiana": "lau",
        "maine": "meu",
        "maryland": "mdu",
        "massachusetts": "mau",
        "michigan": "miu",
        "minnesota": "mnu",
        "mississippi": "msu",
        "missouri": "mou",
        "montana": "mtu",
        "nebraska": "nbu",
        "nevada": "nvu",
        "newhampshire": "nhu",
        "newjersey": "nju",
        "newmexico": "nmu",
        "newyork": "nyu",
        "newyorkstate": "nyu",
        "northcarolina": "ncu",
        "northdakota": "ndu",
        "ohio": "ohu",
        "oklahoma": "oku",
        "oregon": "oru",
        "pennsylvania": "pau",
        "rhodeisland": "riu",
        "southcarolina": "scu",
        "southdakota": "sdu",
        "tennessee": "tnu",
        "texas": "txu",
        "utah": "utu",
        "vermont": "vtu",
        "virginia": "vau",
        "washington": "wau",
        "washingtonstate": "wau",
        "westvirginia": "wvu",
        "wisconsin": "wiu",
        "wyoming": "wyu",

        "australiancapitalterritory": "aca",
        "queensland": "qea",
        "tasmania": "tma",
        "victoria": "vra",
        "westernaustralia": "wea",
        "newsouthwales": "xna",
        "northernterritory": "xoa",
        "southaustralia": "xra",
    }
    place = place_raw.replace(" ", "").lower()
    try:
        result = long_country_codes[place]
    except KeyError as e:
        if len({e}) == 3:
            logger.info(f"Advisory: assuming place name ({e}) has already been processed.")
        else:
            logger.warning(f"Warning: {e} is not a recognised place name; it has been passed on unchanged.")
        result = place_raw
    return result


def get_long_country_code(country: str, place: str) -> str:
    ## USA & UK return a detailed 3-digit code based on local region
    return place.strip().lower() if len(place) == 3 else country


def validate(record: Record) -> bool:
    mandatory = [
        "sublib",
        "langs",
        "title",
        "country",
        "place",
        "publisher",
        "pub_year",
        "extent",
        "size",
        "sale_dates",
        "barcode",
    ]
    for i, field in enumerate(fields(record)):
        name = field.name
        is_valid = True
        if name in mandatory and not getattr(record, name):
            logger.warning(f"Record no. {i} is missing the mandatory field '{name}'.")
            is_valid = False
    return is_valid
    # return record


def norm_dates(raw: str) -> list[str]:
    result = [date.strip() for date in raw.split(",")]
    return result


def norm_size(raw: str) -> int:
    raw = strip_unwanted(r"cm",raw)
    return int(raw)


def norm_pages(pages_raw: str) -> str:
    pages = strip_unwanted(r"pages|\[|\]", pages_raw)
    if "approximately" in pages:
        pages = re.sub(r"\s?approximately\s?", "", pages)
        pages = pages + "?"
    return pages


def norm_year(year_raw: str) -> str:
    year = strip_unwanted(r"[\[\]]", year_raw)
    return year


def strip_unwanted(pattern: str, raw: str) -> str:
  clean = re.sub(pattern, "", raw)
  return clean


def check_for_approx(raw_string: str) -> tuple[str, bool]:
    clean = str(raw_string).strip()
    if clean[-1] == "?":
        is_approx = True
        # raw = trim_mistaken_decimals(raw[:-1].rstrip())
        clean = clean[:-1].rstrip()
    else:
        is_approx = False
    clean = trim_mistaken_decimals(clean)
    return (clean, is_approx)


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string


def create_date_list(dates_raw: str) -> list[str]:
    dates_raw = re.sub(r"\s|\.0", "", dates_raw)
    dates = dates_raw.split(",")
    return dates


def extract_from_excel(excel_sheet) -> list[list[str]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    sheet = []
    for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
        row = []
        if not excel_row[0]:
            break
        for col in excel_row:
            if col:
                data = str(col).strip()
                data = trim_mistaken_decimals(data)
            else:
                data = ""
            row.append(data)
        sheet.append(row)
    return sheet


# def parse_spreadsheet_data(raw_data) -> list[Record]:
#     sheet = extract_from_excel(raw_data)
def parse_spreadsheet(sheet: list[list[str]]) -> list[Record]:
    current_time = datetime.now()
    records = []
    for row in sheet:
        cols = iter(row)
        sublibrary = next(cols)
        langs = norm_langs(next(cols))
        isbn = next(cols)
        title = Title(next(cols), next(cols))
        subtitle = Title(next(cols), next(cols))
        parallel_title = Title(next(cols), next(cols))
        parallel_subtitle = Title(next(cols), next(cols))
        country = norm_country(next(cols))
        place = next(cols)
        publisher = next(cols)
        pub_date, pub_date_is_approx = check_for_approx(norm_year(next(cols)))
        copyright_ = next(cols).replace("©","").strip()
        extent, extent_is_approx = check_for_approx(norm_pages(next(cols)))
        size = norm_size(next(cols))
        series_title = next(cols)
        series_enum = next(cols)
        note = next(cols)
        sale_code = next(cols)
        date_of_sale = create_date_list(next(cols))
        hol_notes = next(cols)
        donation = next(cols)
        barcode = next(cols)

        record = Record(
            sublibrary,
            langs,
            isbn,
            title,
            subtitle,
            parallel_title,
            parallel_subtitle,
            country,
            place,
            publisher,
            pub_date,
            copyright_,
            extent,
            size,
            series_title,
            series_enum,
            note,
            sale_code,
            date_of_sale,
            hol_notes,
            donation,
            barcode,

            pub_date_is_approx,
            extent_is_approx,
            current_time,

            sequence_number = 1,
            links = [880, []],
        )
        validate(record)
        records.append(record)
        # pprint(record: Record) -> Result
    return records


def build_000(record: Record) -> Result:
    """leader (0 is only for sorting purposes; should read 'LDR')"""
    field_num = 0
    i1, i2 = -2, -2  ## "This field has no indicators or subfield codes."
    content = "00000nam a22000003i 4500"
    error = None
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_005(record: Record) -> Result:
    """date & time of transaction
    "The date requires 8 numeric characters in the pattern yyyymmdd. The time requires 8 numeric characters in the pattern hhmmss.f, expressed in terms of the 24-hour (00-23) clock."
    """
    field_num = 5
    i1, i2 = -2, -2  ## "This field has no indicators or subfield codes."
    error = None
    standard_time = record.timestamp.now(timezone.utc)
    ## NB: python produces this format: YYYY-MM-DD HH:MM:SS.ffffff, e.g. 2020-09-30 12:37:55.713351
    timestamp = str(standard_time).translate(str.maketrans("", "", " -:"))[:16]
    success = (field_num, [[i1, i2, timestamp]])
    return Result(success, error)


def build_008(record: Record) -> Result:
    """pub year & main language?"""
    field_num = 8
    i1, i2 = -2, -2  ## "Field has no indicators or subfield codes"
    error = None
    t = record.timestamp
    date_entered_on_file = str(t.year)[2:] + str(t.month).zfill(2) + str(t.day).zfill(2)
    pub_status = "s"
    date_1 = record.pub_year
    date_2 = 4 * "|"
    place_of_pub = record.country.ljust(3, "\\")
    books_configuration = (14*"|") + "\\" + (2*"|")
    lang = record.langs[0].ljust(3, "\\")
    modified_and_cataloging = 2*"|"
    content = f"{date_entered_on_file}{pub_status}{date_1}{date_2}{place_of_pub}{books_configuration}{lang}{modified_and_cataloging}"
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_033(record: Record) -> Result:
    """sales dates"""
    field_num = 33
    i1 = 0 if len(record.sale_dates) == 1 else 1
    i2 = -1
    content = f"$a{"$a".join(record.sale_dates)}"
    error = None
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_040(record: Record) -> Result:
    """catalogued in Oxford"""
    field_num = 40
    i1, i2 = -1, -1
    content = "$aUkOxU$beng$erda$cUkOxU"
    error = None
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_245(record: Record) -> Result:
    """Title
    Field 245 ends with a period, even when another mark of punctuation is present, unless the last word in the field is an abbreviation, initial/letter, or data that ends with final punctuation.
    """
    field_num = 245
    i1 = 0
    nonfiling = 0
    has_chinese_title = bool(record.title.transliteration)
    if has_chinese_title:
        title, subtitle = record.title.transliteration, record.subtitle.transliteration
        chinese_title = combine(record.title.original, record.subtitle.original)
        chinese_title = end_field_with_period(chinese_title)
        # nonfiling = 0
        sequence_number = seq_num(record.sequence_number)
        linkage = f"$6880-{sequence_number}"
        build_880(record, chinese_title, i1, nonfiling, field_num, sequence_number)
    else:
        title, subtitle = record.title.original, record.subtitle.original
        nonfiling, title = check_for_nonfiling(title)
        linkage = ""
    title = combine(title, subtitle) if title else ""
    title = end_field_with_period(title)
    i2 = nonfiling
    if title:
        content = f"{linkage}$a{title}"
        success = (field_num, [[i1, i2, content]])
        error = None
    else:
        success = None
        error = (field_num, "")
    return Result(success, error)


def build_264(record: Record) -> Result:
    """publisher & copyright"""
    field_num = 264
    i1 = -1
    i2 = 1  ## "Publication: Field contains a statement relating to the publication, release, or issuing of a resource."
    # i2 = 0
    content = []
    error = None
    place = f"$a{record.place} "
    publisher = f":$b{record.publisher}"
    pub_year = record.pub_year
    if record.pub_year_is_approx:
        pub_year = f"[{pub_year}?]"
    pub_year = f",$c{pub_year}"
    content.append([i1, i2, f"{place}{publisher}{pub_year}"])
    if record.copyright:
        ## WHY i2=4 ("Copyright notice date") subfield $c ("Date of production, publication, distribution, manufacture, or copyright notice (R)")??
        content.append([i1, -1, f"$a\u00a9 {record.copyright}"])
    success = (field_num, content)
    return Result(success, error)


def build_300(record: Record) -> Result:
    """physical description"""
    field_num = 300
    i1, i2 = -1, -1 ## "undefined"
    error = None
    pages = f"{record.extent} pages"
    if record.extent_is_approx:
        pages = f"approximately {pages}"
    size = f"{record.size} cm"
    content = f"$a{pages} ;$c{size}"
    success = (field_num, [[i1, i2, content]])
    return Result((success), error)


def build_336(record: Record) -> Result:
    """content type (boilerplate)"""
    field_num = 336
    error = None
    content = "$atext$2rdacontent"
    success = (field_num, [[-1, -1,content]])
    return Result(success, error)


def build_337(record: Record) -> Result:
    """media type (boilerplate)"""
    field_num = 337
    i1, i2 = -1, -1
    content = "$aunmediated$2rdamedia"
    error = None
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_338(record: Record) -> Result:
    """carrier type (boilerplate)"""
    field_num = 338
    i1, i2 = -1, -1
    content =  "$avolume$2rdacarrier"
    error = None
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_876(record: Record) -> Result:
    """notes / donations / barcode"""
    field_num = 876
    i1, i2 = -1, -1
    error = None
    notes = record.hol_notes
    # notes = record.notes
    notes = f"$z{notes}" if notes else ""
    donation = f"$z{record.donation}" if record.donation else ""
    barcode = f"$p{record.barcode}" if record.barcode else ""
    content = f"{barcode}{donation}{notes}"
    success = (field_num, [[i1, i2, content]])
    return Result(success, error)


def build_904(record: Record) -> Result:
    """authority (boilerplate)"""
    field_num = 904
    i1, i2 = -1, -1
    error = None
    content =  "$aOxford Local Record"
    success = (field_num, [[i1, i2,content]])
    return Result(success, error)


def build_024(record: Record) -> Result:  ##optional
    """sales code (if exists)"""
    field_num = 24
    i1 = 8
    i2 = -1
    # content = f"$a{record.sales_code}" if record.sales_code else ""
    if record.sales_code:
        content = f"$a{record.sales_code}"
        success = (field_num, [[i1, i2, content]])
        error = None
    else:
        error = (field_num, "")
        success = None
    return Result(success, error)


def build_041(record: Record) -> Result:  ##optional
    """language codes if not monolingual"""
    i1 = -1  ## "No information...as to whether the item is or includes a translation."
    # i1 = 0  ## "No information...as to whether the item is or includes a translation."
    i2 = -1  ## "(followed by) MARC language code"
    field_num = 41
    content = ""
    is_multi_lingual = len(record.langs) > 1
    if is_multi_lingual:
        ## OPTION 1
        # if record.title.transliteration:
        #     main_lang = record.langs[0]
        #     others = record.langs[1:]
        #     lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # else:
        #     lang_list = f"$a{"$a".join(record.lang)}"
        ## OPTION 2
        # main_lang = record.langs[0]
        # others = record.langs[1:]
        # lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # result = (build_field(41, [[i1,i2, lang_list]]))
        ## OPTION 3
        content = f"$a{"$a".join(record.langs)}"
        success = (field_num, [[i1,i2, content]])
        error = None
    else:
        error = (field_num, "")
        success = None
    return Result(success, error)


def build_246(record: Record) -> Result:  ##optional
    """
    Varying Form of Title
    Holds parallel Western title AND/OR original Chinese character title
    NB: Initial articles (e.g., The, La) are generally not recorded in field 246 unless the intent is to file on the article. [https://www.loc.gov/marc/bibliographic/bd246.html]
    """
    field_num = 246
    i1 = 3  ## "No note, added entry"
    i2 = 1  ## parallel title
    has_parallel_title = bool(record.parallel_title.original)
    has_chinese_parallel_title = bool(record.parallel_title.transliteration)
    sequence_number = seq_num(record.sequence_number)
    linkage = f"$6880-{sequence_number}" if has_chinese_parallel_title else ""
    # linkage = "$6880-01" if has_chinese_parallel_title else ""
    if has_chinese_parallel_title:
        parallel_title = record.parallel_title.transliteration
        parallel_subtitle = record.parallel_subtitle.transliteration
        chinese_parallel_title = combine(record.parallel_title.original, record.parallel_subtitle.original)
        # sequence_number = seq_num(record.sequence_number)
        linkage = f"$6880-{sequence_number}"
        build_880(record, chinese_parallel_title,i1, i2, "246", sequence_number)
    elif has_parallel_title:  ## (i.e. Western script)
        parallel_title, parallel_subtitle = record.parallel_title.original, record.parallel_subtitle.original
    else:
        parallel_title, parallel_subtitle = "", ""
    if parallel_title:
        content = linkage + "$a" + combine(parallel_title, parallel_subtitle)
        success = (field_num, [[i1, i2, content]])
        error = None
    else:
        success = None
        error = (field_num, "")
    return Result(success, error)


def build_490(record: Record) -> Result:  ## optional
    """Series Statement"""
    field_num = 490
    i1 = 0  ## Series not traced: No series added entry is desired for the series.
    i2 = -1
    series_title = f"$a{record.series_title}" if record.series_title else ""
    series_enum = f"$v{record.series_enum}" if record.series_enum else ""
    sep = " ;" if series_title and series_enum else ""
    content = series_title + sep + series_enum
    # return build_field(490, [[i1, i2, content ]])
    if content:
        error = None
        success = (field_num, [[i1, i2, content ]])
    else:
        success = None
        error = (field_num, "")
    return Result(success, error)


def build_500(record: Record) -> Result:  ##optional
    """general notes
    Punctuation - Field 500 ends with a period unless another mark of punctuation is present. If the final subfield is subfield $5, the mark of punctuation precedes that subfield.
    """
    # notes = record.hol_notes
    field_num = 500
    i1 = -1
    i2 = -1
    # content = "$a" + end_field_with_period(record.notes) if record.notes else ""
    if record.notes:
        content = "$a" + end_field_with_period(record.notes)
        success = (field_num, [[i1, i2, content]])
        error = None
    else:
        success = None
        error = (field_num, "")
    return Result(success, error)

# TODO: need an item with both a Chinese title and subtitle to test the sequence number logic.
def build_880(record, title, i1, i2, caller, sequence_number) -> None:  ##optional
    """Alternate Graphic Representation
    NB. unlike the other fields, this isn't called directly but by the linked field"""
    record.sequence_number += 1
    content = f"$6{field_prefix(caller)}-{sequence_number}$a{title}"
    record.links[1].append(build_line(line_prefix(880), i1, i2, content))


def build_field(result: Result) -> tuple[int, list[str]] | None:
    """
    silently suppresses optional fields if empty;
    stops with error if required field is empty
    """
    optional_fields = [
        24,     ## sales code
        41,     ## language if not monolingual
        246,    ## parallel title
        490,    ## series statement
        500,    ## general notes
        880,    ## transliteration
    ]
    output: tuple[int, list[str]] | None = None
    if result.is_err:
        numeric_tag, error_msg = result.is_err
        if numeric_tag not in optional_fields:
            msg = error_msg if error_msg else f"Data for required field {str(numeric_tag).zfill(3)} is required."
            logger.warning(msg)
            raise MissingFieldError(msg)
            # logger.warning(f"Data for required field {str(numeric_tag).zfill(3)} is required.")
            # raise MissingFieldError(f"Data for required field {str(numeric_tag).zfill(3)} is required.")
    elif result.is_ok:
        numeric_tag, data = result.is_ok
        lines = []
        for line in data:
            i1 = expand_indicators(line[0])
            i2 = expand_indicators(line[1])
            content = line[2]
            if content:
                lines.append(build_line(line_prefix(numeric_tag), i1, i2, content))
            else:
                break
        output = (numeric_tag, lines) if lines else None
    return output


def build_line(line_start, i1, i2, content) -> str:
    return (f"{line_start}{i1}{i2}{content}")


def line_prefix(numeric_tag: int) -> str:
    display_tag = "LDR" if numeric_tag == 0 else seq_num(numeric_tag)
    return f"={field_prefix(display_tag)}  "


def field_prefix(field_number: str) -> str:
    return str(field_number).zfill(3)


def seq_num(sequence_number: int) -> str:
    return str(sequence_number).zfill(2)


def expand_indicators(indicator: int) -> str:
    expansions: dict[int, str] = {-2: "", -1: "\\"}
    expansion: str = expansions.get(indicator, str(indicator))
    return expansion


def combine(title: str, subtitle: str, sep: str=" :$b") -> str:
    if subtitle:
        title += sep + subtitle
    return title


def end_field_with_period(text: str) -> str:
    text = text.strip()
    if text and text[-1] not in "?!.":
        text = text + "."
    return text


def check_for_nonfiling(title: str, lang: str="eng") -> tuple[int,str]:
    """
    Check for manual nonfiling indicator (@@) & returns its position + extracts it from title string
    if no manual indication, check for nonfiling words according to language of title"""
    nonfiling_words = {
        "eng": ("the", "a", "an"),
        "fr": ("le", "la", "les", "l'", "un", "une"),
        "it": ("lo", "il", "i", "l'", "gli", "le", "un", "una", "un'"),
        "sp": ("el", "la", "las", "los", "un", "una", "unos", "unas"),
        "gw": ("der", "die", "das", "ein", "eine"),
        "ne": ("de", "het"),
        "sw": ("en", "ett", "den", "det", "de"),
        "dk": ("en", "et"),
        "no": ("en", "ei", "et"),
        "po": ("o", "a", "os", "as", "um", "uma", "uns", "umas"),
        "chi": (),
    }
    break_char = "@@"
    nonfiling = title.find(break_char)
    result = (0, title)
    if nonfiling > 0:
        result = (nonfiling, title.replace(break_char, "", 1))
    else:
        for article in nonfiling_words[lang]:
            test = re.match(f"({article}\\s?[^\\w\\s]?)\\w", title, re.I)
            if test:
                result = (test.span()[1] - 1, title)
                break
    return result


def build_mark_records(records: list[Record]) -> list[list]:
    mark_records: list = []
    for record in records:
        mark_record: list = []
        for builder in (
            build_000,
            build_040,
            build_336,
            build_337,
            build_338,
            build_904,
            build_005,
            build_008,
            build_033,
            build_245,
            build_264,
            build_300,
            build_490,
            build_876,
            build_024,
            build_041,
            build_246,
            build_500,
        ):
            field = build_field(builder(record))
            if field:
                mark_record.append(field)
        if record.links:
            mark_record.append(record.links)
        mark_record.sort(key=lambda field: field[0])
        mark_records.append(mark_record)
    return mark_records


def write_mrk_file(data, file_name: str="output.mrk") -> None:
    mrk_file_dir = Path("mrk_files")
    if not mrk_file_dir.is_dir():
        mrk_file_dir.mkdir()
        try:
            mrk_file_dir.mkdir()
            logger.info(f"Directory '{mrk_file_dir}' created successfully.")
        except FileExistsError:
            logger.info(f"Directory '{mrk_file_dir}' already exists.")
        except PermissionError:
            logger.warning(f"Permission denied: Unable to create '{mrk_file_dir}'.")
        except Exception as e:
            logger.warning(f"An error occurred: {e}")
    out_file = mrk_file_dir / file_name
    with open(out_file, "w", encoding="utf-8") as f:
        for record in data:
            for field in record:
                for line in field[1]:
                    # print(line)
                    f.write(line)
                    f.write("\n")
            f.write("\n")


def get_records(excel_file_address: Path) -> list[Record]:
    excel_file_name = str(excel_file_address.resolve())
    worksheet = load_workbook(filename=excel_file_name).active
    data = extract_from_excel(worksheet)
    records = parse_spreadsheet(data)
    return records


def process_excel_file(excel_file_address: Path) -> None:
    records = get_records(excel_file_address)
    mark_record_set = build_mark_records(records)
    write_mrk_file(mark_record_set, f"{excel_file_address.stem}.paul.mrk")
    # pprint(records)
    # for count, record in enumerate(records):
    #     print(f">> {count}")
    #     pprint(record: Record) -> Result


def main() -> None:
    # process_excel_file(sys.argv[1])
    for file in Path("excel_files").glob("*.xlsx"):
        logger.info(f"\n>>>>> processing: {file.name}")
        print(f">>>>> processing: {file.name}")
        process_excel_file(file)


if __name__ == "__main__":
    main()
