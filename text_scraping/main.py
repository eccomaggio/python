from pathlib import Path
from csv import reader, writer
import argparse
from pprint import pprint
from openpyxl import load_workbook  # type: ignore
import datetime
import pytz



def argument_parser() -> tuple[Path, Path, Path]:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Process some text files.")
    parser.add_argument("-s", "--source", type=Path, default=Path("input.txt"), help="Source file path")
    parser.add_argument("-o", "--output", type=Path, default=Path("output.csv"), help="Destination file path")
    parser.add_argument("-c", "--concordance", type=Path, default=Path("concordance.xlsx"), help="Concordance file path")
    args = parser.parse_args()
    return (args.source, args.output, args.concordance)


def read_lines(file_path: Path) -> list[str]:
    """Read lines from a file and return them as a list."""
    with file_path.open("r", encoding="utf-8") as file:
        return file.readlines()


def write_lines(file_path: Path, lines: list[str]) -> None:
    """Write a list of lines to a file."""
    with file_path.open("w", encoding="utf-8") as file:
        file.writelines(lines)


def write_csv(file_path: Path, data: list) -> None:
    """Write a list of lists to a CSV file."""
    with file_path.open("w", encoding="utf-8", newline='') as file:
        csv_writer = writer(file)
        csv_writer.writerows(data)


def read_csv(file_path: Path) -> list:
    """Read a CSV file and return its content as a list of lists."""
    with file_path.open("r", encoding="utf-8") as file:
        csv_reader = reader(file)
        return list(csv_reader)


def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    excel_file_name: str = str(excel_file_address.resolve())
    excel_sheet = load_workbook(filename=excel_file_name).active
    sheet = []
    for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
        row = []
        if not excel_row[0]:
            break
        for col in excel_row:
            if col:
                data = str(col).strip()
                data = trim_mistaken_decimals(data)
            else:
                data = ""
            row.append(data)
        sheet.append(row)
    return sheet


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string

def normalise_concordance(raw: list[list[str]]) -> dict[int, list[str]]:
    concordance: dict[int, list[str]] = {}
    for row in raw:
        object_id = row[0]
        object_num = row[1]
        if row[5].isnumeric():
            concordance[int(row[5])] = [object_id, object_num]
    return concordance


def make_concordance(excel_file_path: Path) -> dict[int, list[str]]:
    raw = extract_from_excel(Path(excel_file_path))
    concordance = normalise_concordance(raw)
    return concordance


def group_lines(raw_lines: list) -> dict[int, list]:
    """Process the text from the source list and return a list of processed lines."""
    processed_text: dict[int, list[str]] = {}
    section_number:int = 0
    the_rest: list[int] = []
    is_multiple = False
    section: list[str] = []
    for line in raw_lines:
        line = line.strip()
        if not line:
            continue
        ## If the line starts with @@number, it indicates a new section
        if (command_from_here := line.find("@@")) > -1:
            line = line[command_from_here + 2:]
            if not line[0].isnumeric():
                # ignore any instructions that aren't numbers for now
                continue
            # If there is an existing section, save it
            # @@1&2 means section 2 is a repeat of section 1,
            # so add it to the processed text
            print(f"Processing new section: {line}")
            if section:
                processed_text[section_number] = section
            if is_multiple:
                for part in [the_rest]:
                    print(f"Adding section {section_number} to processed text for part {part}")
                    processed_text[part[0]] = section
            section = []
            is_multiple = False

            ## Process the new section number
            tmp: list[int] = [int(part) for part in line.split("&")]
            section_number = tmp[0]
            the_rest = tmp[1:]
            is_multiple = len(the_rest) > 0
            # ignore all entries above 50 (for now!)
            if section_number > 50:
                break
        else:
            section.append(line)
    if section:
        processed_text[section_number] = section
    return processed_text


def prepare_for_csv(processed_text: dict[int, list], concordance: dict[int, list]) -> list[tuple]:
    # ID (int)	Import identifier	Audience ("public")	Date	Notes	Reason	Sort [int]	Source	Status ("05 Published")	Text	Title / Ref. No.	Type ("catalogue text")	Language ("en") -> [=13]
    headings = ("ID", "Import identifier", "Audience", "Date", "Notes", "Reason", "Sort", "Source", "Status", "Text", "Title / Ref. No.", "Type", "Language")

    output: list[tuple[int|str, str, str, str, str, str, int|str, str, str, str, str, str, str]] = []
    import_id: str
    audience = "public"
    date = str(datetime.datetime.now(pytz.timezone("Europe/London")))[:10]
    notes = ""
    reason = ""
    source = ""
    status = "05 Published"
    _type = "catalogue text"
    language = "en"
    output.append(headings)
    for num, lines in processed_text.items():
        if not (tmp := concordance.get(num, "")):
            continue
        else:
            import_id = tmp[0]
        sort = 1
        text = "\n".join(lines)
        ref_no = ""
        output.append((num, import_id, audience, date, notes, reason, sort, source, status, text, ref_no, _type, language))
    return output




def main() -> None:
    (source_file,
     destination_file,
     concordance_file) = argument_parser()
    concordance = make_concordance(Path(concordance_file))
    print(f"Reading from {source_file.name} and writing to {destination_file.name}...")
    raw_lines: list[str] = read_lines(source_file)
    processed_text = group_lines(raw_lines)
    del raw_lines
    csv_ready_text = prepare_for_csv(processed_text, concordance)
    del processed_text

    print(f"Processed {len(csv_ready_text) - 1} sections from {source_file.name}." )
    # pprint(processed_text.get(1, []))
    # pprint(concordance)
    # write_csv(destination_file, [[num, "\\n".join(lines)] for num, lines in processed_text.items()])
    write_csv(destination_file, csv_ready_text)



if __name__ == "__main__":
    main()
